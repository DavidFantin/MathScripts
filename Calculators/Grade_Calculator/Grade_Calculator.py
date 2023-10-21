#David Fantin
#2/21/20
#Purpose: To calculate a total grade based on weighted percentages

import openpyxl as opx
import ctypes

ctypes.windll.kernel32.SetConsoleTitleW("Grade Calculator")

# Where all of the calculations are done
def calc(r, scoreCol, weightCol, TOTAL):   
    q_score = sheet.cell(row = r, column = scoreCol).value
    q_weight = sheet.cell(row = r, column = weightCol).value
    result = (q_weight/100) * q_score
    TOTAL = TOTAL + result
    return TOTAL

while True:
    loc = "C:\\Users\\dfant\\Documents\\Grade_Calculator.xlsx"
    wb = opx.load_workbook(loc)
    myfile = wb
    sheet = wb[wb.sheetnames[0]]

    #Main Program
    x = 3 #Starting Row
    scoreCol = 3 #Score Column
    weightCol = 2 #Weight % Column
    TOTAL = 0 #Final Answer
    
    #Calculates TOTAL
    while sheet.cell(row = x, column = scoreCol).value != None:
        TOTAL = calc(x, scoreCol, weightCol, TOTAL)
        x = x+1

    TOTAL = TOTAL

    #Prints TOTAL
    print("<#><#><#><#><#><#><#>\n\nThe total score is:")
    print(TOTAL)
    print("____________________")

    Decision = input("1. Press ENTER to recalculate\n2. Enter '1' to write the total to the file\n>>>")

    if Decision == '1':

        sheet.cell(row = 3, column = 4).value = TOTAL
        
        while True:
            try: #Checks if Excel is OPEN or NOT OPEN
                myfile = open(loc, "r+")
            except IOError: #If OPEN
                print ("***********************************************")
                print ("Could not save to the file! Please close Excel!")
                print ("***********************************************")
                print ("Press ENTER to try again...")
                print ("***************************")
                input()
                continue
            break
        with myfile: #Writes TOTAL to the Excel file
            wb.save(loc)
            input("Write Successfull!! Press ENTER to End Program...")
            break
    elif Decision == '1701':
        input("___________________          _-_\n\==============_=_/ ____.---'---`---.____\n            \_ \    \----._________.----/\n              \ \   /  /    `-_-'\n          __,--`.`-'..'-_\n         /____          ||\n              `--.____,-'\n\nLive long and prosper!")
        break
    else: #If NOT OPEN
        continue
